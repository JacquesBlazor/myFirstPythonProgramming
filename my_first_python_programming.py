import time
from time import localtime, gmtime, strftime, sleep
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import json
from PIL import Image
from io import BytesIO
import sys
import pyocr
import pyocr.builders
import cv2
import pytesseract
import numpy as np
from time import sleep
import openpyxl
from gspread import Client
import shutil
from authlib.client import AssertionSession
import requests
import os
import boto3

def create_assertion_session(conf_file, gss_scopes, subject=None):
    with open(conf_file, 'r') as f:
        conf = json.load(f)
    token_url = conf['token_uri']
    issuer = conf['client_email']
    key = conf['private_key']
    key_id = conf.get('private_key_id')
    header = {'alg': 'RS256'}
    if key_id:
        header['kid'] = key_id
    # Google puts scope in payload
    claims = {'scope': ' '.join(gss_scopes)}
    return AssertionSession(
        grant_type=AssertionSession.JWT_BEARER_GRANT_TYPE,
        token_url=token_url,
        issuer=issuer,
        audience=token_url,
        claims=claims,
        subject=subject,
        key=key,
        header=header,
    )
gss_scopes = [
    'https://spreadsheets.google.com/feeds',
    'https://www.googleapis.com/auth/drive',
]

def slackPrint(slackMessage):
        # 使用 Slack 發送訊息更新近況
        slackUrl = 'https://hooks.slack.com/services/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'
        slackDictHeaders = {'Content-type': 'application/json'}
        pydictPayload = { }
        pydictPayload["text"] = slackMessage
        jsonPayload = json.dumps(pydictPayload)
        slackRtn = requests.post(slackUrl , data=jsonPayload, headers=slackDictHeaders)
        print('update to slack@', slackMessage)

# # 程式開始, 設定存圖檔案環境變數
captchaDlPath = 'captchaDlPath//'
convertedImgPath = 'convertedImgPath//'
# 廠商基本資料查詢網頁 URL
expRegDataUrl = 'https://get.vendors_info.com/index.asp'  # 已修改掉真正的網址

# 將本機現有 excel 資料表移至另個資料夾並更改名稱為現在時間做為備份, 然後重新下載最近的檔案開始運作
if os.path.exists('xls_document.xlsx') and os.path.isfile('xls_document.xlsx'):
    shutil.move('xls_document.xlsx', './/xlsotherfiles/'+strftime('xls_%Y%m%d%H%M%S', localtime()) + '.xlsx' )
gcpXLsheetUrl = 'https://docs.google.com/spreadsheets/d/1FrLtAoxf1DJpGwzn1FkTeQSHEA-TmrFlze0OdkVeWi4/export?format=xlsx&id=1FrLtAoxf1DJpGwzn1FkTeQSHEA-TmrFlze0OdkVeWi4'
excelRequestedData= requests.get(gcpXLsheetUrl)
with open('xls_document.xlsx', 'wb+') as exlFile2beSaved:
        exlFile2beSaved.write(excelRequestedData.content)

# 開啟本機 excel 資料表
xlWbook =openpyxl.load_workbook(r'xls_document.xlsx')
xlSheet = xlWbook.worksheets[0]

# 找出 Excel Sheet 資料表中欄位項目 行列 1 表示欄位名稱的清單
allTitleFields = []
for i in range(1, xlSheet.max_column+1):
    allTitleFields.extend([xlSheet.cell(row=1, column=i).value])

indexNumLists = []  # 找出此次要繼續尋找資料的列號 indexNumList 清單
for i in range(8563, 9000):
    if xlSheet.cell(row=i, column=1).value == None and xlSheet.cell(row=i, column=5).value == None and xlSheet.cell(row=i, column=6).value == None:
        indexNumLists.extend([i])
totalWorkItems = len(indexNumLists)  # 找出此次要繼續尋找資料的數量

if totalWorkItems == 0:
    # 看來沒有要繼續工作的資料, 程式結束
    print('Program ends. The excle file has full data which does not require to further process it. ')
    slackPrint('|Program ends|')
    sys.exti(1)

# 開始連線開啟 Google Sheet 資料表
gss_session = create_assertion_session('MyPythonStock-0a059f9673c0.json', gss_scopes)
gss_client = Client(None, gss_session)
gcpSheet = gss_client.open_by_key('1FrLtAoxf1DJpGwzn1FkTeQSHEA-TmrFlze0OdkVeWi4').sheet1


# 開始繼續工作直到沒有要作業的項目
loopEnd = 0
slackCounter = 0
averagedItems = []
while totalWorkItems != 0:
    tryCount = 0  # 變數 tryCount 表示已查詢過的次數
    flag2RetryEngNames = True    # 變數 flag2RetryEngNames 表示此次要不要換個英文名字再重新查詢資料
    flagDataFound = False  # 變數 flagDataFound 表示查詢資料結果是否成功
    codeIsIncorrect = True    # 判斷驗證碼是否輸入正確, 它應該是有 6 碼, 並且只有小寫英文或數字
    engNameLessThen2  = False  # 用來判斷目前剩下的英文字少於 2 個單字

    # 開啟 Chrome 瀏覽器
    agentChrome = webdriver.Chrome()
    print("Current session is {}".format(agentChrome.session_id))
    agentChrome.maximize_window()
    agentChrome.get(expRegDataUrl)

    # 拿出第一筆資料的列號, 然後將 totalWorkItems 重設為剩下筆數的數量
    if not loopEnd == 0:
        loopAverage = int(loopEnd-loopStart)
        averagedItems.append(loopAverage)
    else:
        loopAverage = 0
    indexNumber = indexNumLists.pop()
    totalWorkItems = len(indexNumLists)
    print('Now working on row: ', indexNumber)
    loopStart = time.time()

    slackCounter += 1
    # 使用 Slack 發送訊息, 每10發
    if (slackCounter % 5) == 1:
        if slackCounter != 1:
            totalAverage = sum(averagedItems)//len(averagedItems)
        else:
            totalAverage = 0
        slackPrint('#'+str(slackCounter)+'a|'+'@'+str(indexNumber)+'|'+str(totalAverage)+'s|'+strftime('%d'+'/'+'%H'+':'+'%M'+':'+'%S', localtime()))
    elif (loopAverage > 350):
        slackPrint('*#'+str(slackCounter)+'a|'+'@'+str(indexNumber)+'|'+str(loopAverage)+'s|'+strftime('%d'+'/'+'%H'+':'+'%M'+':'+'%S', localtime()))
        # setting the rule to start the instance to be started after 5 mins when I stop it later.
        myruleName = 'StartEC2Instances'
        schedEvent = boto3.client('events')
        currentEvent = schedEvent.describe_rule(Name=myruleName)
        currentExpression = currentEvent['ScheduleExpression']
        currentHtime = int(strftime('%H', gmtime()))
        currentMtime = int(strftime('%M', gmtime()))
        increaseHtime =  ( int(currentMtime)+3 ) //  60
        targetHtime  = '%02d' % (( currentHtime + increaseHtime ) % 24 )
        targetMtime = '%02d' % ((int(currentMtime)+3) % 60 )
        proposedScheduleExpression = 'cron('+targetMtime + ' ' + targetHtime + ' * * ? *)'
        schedEvent.put_rule(Name=myruleName, ScheduleExpression=proposedScheduleExpression)
        # stop the instance and it should be started after 5 mins by the rule setting above.
        ec2Client = boto3.client('ec2', region_name='us-east-1')
        ec2Response = ec2Client.stop_instances(InstanceIds=['i-0d7225ccfb8de48e1',],)

    # 讀取英文名稱欄位
    companyEngWord = xlSheet.cell(row=indexNumber, column=2).value # 變數 companyEngWord 是要填進欄位 'english_name' 的資料
    companyEngWords = xlSheet.cell(row=indexNumber, column=2).value.split(' ')  # 變數 companyEngWords 是將英文名字拆成單字

    # 欄位 D  更新記錄此時程式處理這筆資料的時間為目前處理的時間
    gcpSheet.update_cell(indexNumber, 4, strftime('%Y%m%d%H%M%S', localtime()))
    xlSheet.cell(row=indexNumber, column=4, value=strftime('%Y%m%d%H%M%S', localtime()))
    sleep(1)

    # 如果還沒找到資料, 或英文字至少有2個單字
    while not flagDataFound or not engNameLessThen2:
         # 當驗證碼不正確或需要再用相同資料改變英文字數嘗試時
        while flag2RetryEngNames or codeIsIncorrect:
             # 使用第一次英文名稱查詢資料
            wordsCount = len(companyEngWords)    # 變數  wordsCount 是看公司英文名字總共有幾個單字
            expRegDataEngName = agentChrome.find_element_by_name('english_name')
            print("Enter: '"+companyEngWord+"'")
            expRegDataEngName.clear()
            expRegDataEngName.send_keys(companyEngWord.rstrip())
            sleep(1)
            # 從 screenshot 中取得驗證碼的圖片
            pngImgName = strftime('img_%Y%m%d%H%M%S', localtime()) + '.png'
            imgName = captchaDlPath + pngImgName
            elmtValid = agentChrome.find_element_by_id('valid')
            elmtValidLeft = elmtValid.location['x']
            elmtValidTop = elmtValid.location['y']
            elmtValidRight = elmtValid.location['x'] + elmtValid.size['width']
            elmtValidBottom = elmtValid.location['y'] + elmtValid.size['height']
            fullPngImg = agentChrome.get_screenshot_as_png()
            captchaImg = Image.open(BytesIO(fullPngImg))
            captchaImg = captchaImg.crop((elmtValidLeft ,elmtValidTop, elmtValidRight, elmtValidBottom))
            captchaImg.save(imgName)
            # 從檔案讀取抓取下來的驗證碼圖片並做圖形處理
            ocrImage = cv2.imread(imgName)
            ocrImage  = cv2.resize(ocrImage, None, fx=3, fy=3, interpolation=cv2.INTER_LINEAR)  # 放大3倍
            ocrImage = cv2.cvtColor(ocrImage, cv2.COLOR_BGR2GRAY)  # 轉灰階
            ret,ocrImage = cv2.threshold(ocrImage,127,255,cv2.THRESH_BINARY)  # 轉黑白
            ocrImage = cv2.morphologyEx(ocrImage, cv2.MORPH_OPEN, np.ones((4,4),np.uint8))  # 去雜訊
            ocrImage = cv2.morphologyEx(ocrImage, cv2.MORPH_CLOSE, np.ones((4,4),np.uint8))  # 補空白點
            # 做完圖形處理驗證碼圖片後送進 OCR 做判讀
            dstImageName=imgName.replace(captchaDlPath, convertedImgPath)
            cv2.imwrite(dstImageName, ocrImage)
            ocrTools = pyocr.get_available_tools()
            if len(ocrTools) == 0:
                print("No OCR tool found")
                sys.exit(1)
            ocrTool =  ocrTools[0]
            ocrText = ocrTool.image_to_string(Image.open(dstImageName), builder=pyocr.builders.TextBuilder())
            # 列印 OCR 判讀後的字
            print("The ocr decode: '", ocrText, "'")
            # 將 OCR 判讀後的字改為只有英文及數字
            ocrIsAlnumText = ''.join(e for e in ocrText if e.isalnum()).lower()
            print("After isalnum process: '",ocrIsAlnumText, "'")
            # 如果英文及數字總合不是6個字元表示判讀錯誤, 需重新產生新的驗證碼再判讀一次
            if len(ocrIsAlnumText) == 6:
                # 如果英文及數字是6個字元表示判讀完畢, 輸入後看是否正確
                codeIsIncorrect = False
                textCheckCode = agentChrome.find_element_by_name('txtCheckCode')
                # 輸入 OCR 處理過的 ocrIsAlnumText 文字
                textCheckCode.send_keys(ocrIsAlnumText)    # 改放OCR後處理過的ocrIsAlnumText文字
                sleep(1)
                agentChrome.find_element_by_name('query').click()
                # 輸入驗證馬後, 判斷結果有沒有錯誤訊息
                try:
                    # 確定有彈出錯誤訊息 ?
                    WebDriverWait(agentChrome, 3).until(EC.alert_is_present(), 'Timed out waiting for alerts to appear')
                    alertNoDataFound = agentChrome.switch_to.alert
                except:
                    # 沒有彈出錯誤訊息, 表示驗證碼正確並查詢有結果  所以有找到資料, 開始蒐集資料內容
                    print('Great! The verification code is correct and The data is found!')
                    flag2RetryEngNames = False
                    flagDataFound = True
                    # 將正確的驗證碼做為檔案名稱, 更改原先寫入的抓取圖和轉檔
                    cvtKeyinfileName = convertedImgPath + ocrIsAlnumText + '.jpg'
                    shutil.move(dstImageName, cvtKeyinfileName)
                    srcKeyinfileName = cvtKeyinfileName.replace(convertedImgPath, captchaDlPath)
                    shutil.move(imgName, srcKeyinfileName)
                    break   ##  有搜尋到資料, 跳出 flag2RetryEngNames or codeIsIncorrect 的迴圈開始蒐集資料內容
                else:
                    print('Alert Text: ', alertNoDataFound.text)
                    if alertNoDataFound.text == '驗證碼輸入錯誤，請重新查詢!!':
                        # alert: "驗證碼輸入錯誤，需重新查詢"
                        print("Code is incorrect. ")
                        codeIsIncorrect = True
                        print("We will loop another verification code and try again. ")
                        sleep(2)
                        alertNoDataFound.accept()
                        print('The alert 驗證碼輸入錯誤，請重新查詢 is accepted!')
                        sleep(2)
                    elif alertNoDataFound.text == '找不到所查詢的資料，請重新查詢!!':
                        # alert: "找不到所查詢的資料，請重新查詢"
                        tryCount += 1
                        print("The inqury data was not found. This is the ", tryCount, " try.")
                        flagDataFound  = False
                        print("We will loop another date and try again. ")
                        sleep(2)
                        alertNoDataFound.accept()
                        sleep(2)
                        print('The alert 找不到所查詢的資料，請重新查詢 is accepted!')
                        gcpSheet.update_cell(indexNumber, 5, 'No')   # 更新欄位 E 表示目前沒有找到資料
                        xlSheet.cell(row=indexNumber, column=5, value= 'No')
                        sleep(1)
                        gcpSheet.update_cell(indexNumber, 7, strftime('%Y%m%d%H%M%S', localtime()))     # 欄位 G  更新記錄此時程式處理這筆資料的時間
                        xlSheet.cell(row=indexNumber, column=7, value=strftime('%Y%m%d%H%M%S', localtime()))
                        sleep(1)
                        gcpSheet.update_cell(indexNumber, 8, tryCount)   # 更新欄位 H 表示做了第幾次嘗試
                        xlSheet.cell(row=indexNumber, column=8, value=tryCount)
                        sleep(1)
                        companyEngWords.pop()    # 把公司的英文名字最後一個單字拿掉.
                        if len(companyEngWords) <2:
                            # 欄位無法接受少於2個英文字的查詢, 此筆資料不需再嘗試
                            flag2RetryEngNames = False
                            flagDataFound = False
                            engNameLessThen2 = True
                            #  這筆資料找不到資料, 將記紀寫回 Google Sheet 資料表
                            gcpSheet.update_cell(indexNumber, 6, tryCount)   # 更新欄位 F 表示已做了所有的嘗試但仍找不到資料
                            xlSheet.cell(row=indexNumber, column=6, value=tryCount)
                            sleep(1)
                            xlWbook.save('xls_document.xlsx')
                            loopEnd = time.time()
                            break   # 跳出 flag2RetryEngNames or codeIsIncorrect 的迴圈. 且資料也沒找到
                        else:
                            # 把公司的英文名字最後一個單字拿掉後再做一次嘗試
                            print("Let's make the name shorter and try again.")
                            engNameLessThen2 = False
                            flag2RetryEngNames = True
                            flagDataFound = False
                            tryThisName = ''
                            for i in range(0, len(companyEngWords)):
                                  tryThisName += str(companyEngWords[i]) + ' '
                            companyEngWord = tryThisName
                            print("I will try this name"+"'",companyEngWord, " next time.")
            else:
                codeIsIncorrect = True
                flag2RetryEngNames = False
                sleep(1)
                agentChrome.find_element_by_name('btnReGenCode').click()
                print('The decode failed. Let\'s regenerate the new code and try again')
        if not flagDataFound:
            #  這筆資料找不到資料, 將記紀寫回 Google Sheet 資料表
            gcpSheet.update_cell(indexNumber, 6, tryCount)   # 更新欄位 F 表示已做了所有的嘗試但仍找不到資料
            xlSheet.cell(row=indexNumber, column=6, value=tryCount)
            sleep(1)
            xlWbook.save('xls_document.xlsx')
            agentChrome.quit()
            loopEnd = time.time()
            break
        else:
            # 找到資料這筆資料, 將第1頁資料寫回 Google Sheet 資料表
            gcpSheet.update_cell(indexNumber, 5, 'Yes')   # 欄位 E 表示有找到資料
            xlSheet.cell(row=indexNumber, column=5, value='Yes')
            sleep(1)
            foundPage1Field = agentChrome.find_elements_by_class_name('td1bg1')  # 欄位清單
            foundPage1Data = agentChrome.find_elements_by_class_name('td2bg1') # 資料清單
            gcpSheet.update_cell(indexNumber, 9, int(len(foundPage1Data)/3)) # 資料個數
            xlSheet.cell(row=indexNumber, column=9, value=len(foundPage1Field))
            sleep(1)
            for i in range(9, 12):    # 資料欄位從 J 開始, J=10, 所以索引是9, 到最後一筆總共的欄位長度, 第一頁只到欄位 L
                for j in range (0, len(foundPage1Field)):
                    if allTitleFields[i] in foundPage1Field[j].text:
                        gcpSheet.update_cell(indexNumber, i+1,  foundPage1Data[j].text)
                        xlSheet.cell(row=indexNumber, column=i+1, value=foundPage1Data[j].text)
                        sleep(1)
                        if i == 9:
                            gcpSheet.update_cell(indexNumber, 1,  foundPage1Data[j].text)  # 資料欄位 A 更新為中文名稱
                            xlSheet.cell(row=indexNumber, column=1, value=foundPage1Data[j].text)
                            sleep(1)
            try:
                foundPage1Link = agentChrome.find_element_by_xpath("//a") # 有超連結
            except:
                print('pleae deal with the exception found!')
            else:
                parentWindowHandle = agentChrome.current_window_handle  #  get handle name
                if foundPage1Link.get_attribute('href')  != '':
                    # 這筆資料還有第2頁, 點開超連結, 將第2頁資料寫回 Google Sheet 資料表
                    print('The 2nd data page found, start 2nd page')
                    foundPage1Link.click()
                    WebDriverWait(agentChrome, 10).until(EC.number_of_windows_to_be(2))
                    allWindowHandles  = agentChrome.window_handles  # get all handle names
                    print('Number of Windows: ',len(allWindowHandles))
                    # switch to new window: if the window is not current, switch to that window
                    for thisWindow in allWindowHandles:
                        if thisWindow != parentWindowHandle:
                             agentChrome.switch_to.window(thisWindow)
                             agentChrome.maximize_window()
                    # 將第2頁資料寫回 Google Sheet 資料表
                    foundPage2FieldnData = agentChrome.find_elements_by_tag_name('td')  # 所有欄位清單
                    for i in range(12, len(allTitleFields)):
                        for j in range (0, len(foundPage2FieldnData)):
                            if allTitleFields[i] in foundPage2FieldnData[j].text and foundPage2FieldnData[j].text != '':
                                if i < 29 :
                                    gcpSheet.update_cell(indexNumber, i+1,  foundPage2FieldnData[j+1].text)
                                    xlSheet.cell(row=indexNumber, column=i+1, value=foundPage2FieldnData[j+1].text)
                                    sleep(1)
                                elif i >= 29:
                                    gcpSheet.update_cell(indexNumber, i+1,  foundPage2FieldnData[j+2].text.replace('\nYes',''))  # Import Qualification=29,Export QualificationExport Qualification=30
                                    xlSheet.cell(row=indexNumber, column=i+1, value=foundPage2FieldnData[j+2].text.replace('\nYes',''))
                                    sleep(1)
                    agentChrome.close()
                    print('Window 2 close')
                    agentChrome.switch_to.window(parentWindowHandle)
            agentChrome.close()
            print('Window 1 close')
            gcpSheet.update_cell(indexNumber, 7, strftime('%Y%m%d%H%M%S', localtime()))    # 欄位 G 更新記錄此時程式更新完這筆資料的時間
            sleep(1)
            xlSheet.cell(row=indexNumber, column=7, value=strftime('%Y%m%d%H%M%S', localtime()))
            print('Google sheet row number: ', indexNumber,' was updated!')
            xlWbook.save('xls_document.xlsx')
            agentChrome.quit()
            loopEnd = time.time()
            break